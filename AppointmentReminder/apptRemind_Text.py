import os
import csv
import glob
import xlrd
import os.path
import sys
import datetime
import time as tm
from pprint import pprint
import openpyxl
import re
import logging
from references_Text import *
from emaillog import emaillog
import redcap
from filePaths import *
import copy
import tkinter



def one(rerun=False):

    # filemode = 'w' if you want to overwrite each previous days log
    # filemode = 'a' if you want a continuous log
    print('Program start')
    logging.basicConfig(filemode = 'a', format='%(levelname)s:%(asctime)s:%(message)s', datefmt='%m/%d/%Y %I:%M:%S %p', filename= logs + 'system.log', level=logging.INFO)

    if rerun == True:
        logging.info('\n')
        logging.info('****************  Program Start  (RERUN MODE)  ******************')
        window = tkinter.Tk()
        window.title('Big Pink Bus Appointment Reminder')
        window.geometry("650x200")
        lbl = tkinter.Label(window, text="Make sure to change invitation settings in Redcap before you continue.\n\nRemember this method of running the program will send notifications to\npatients with an appointment tomorrow, so the reminder needs to go out today."  , font=("Helvetica", 12))

        def ok():
            window.quit()

        def cancel():
            logging.info('Rerun mode cancelled')
            sys.exit(0)

        okbtn = tkinter.Button(window, height=2, width=10, text="Continue", command=ok)
        nobtn = tkinter.Button(window, height=2, width=10, text="Cancel", command=cancel)

        lbl.pack(side=tkinter.TOP, padx=20, pady=20)
        okbtn.pack(side=tkinter.LEFT, padx=130, pady=4)
        nobtn.pack(side=tkinter.RIGHT, padx=130, pady=4)

        window.mainloop()

    else:
        logging.info('\n')
        logging.info('****************  Program Start  ******************')

    # (IDnumber) Load the source workbook and log workbook and identify the next unique ID number from the log workbook. All filepaths are in the seperate filePaths.py file
    try:
        logbook = xlrd.open_workbook(os.path.join(logs, outputlog))
        logsheet = logbook.sheet_by_index(0)
        locationsbook = xlrd.open_workbook(os.path.join(eventLocations, loc))
        sheet_names = locationsbook.sheet_names()
        IDnumber = int(logsheet.cell_value(logsheet.nrows - 1, 0)) + 1
        logbook.release_resources()

    except:
        logging.error('Error locating one of the input workbooks.')
        title = 'ACTION REQUIRED: Error in the Automated Text Reminder System!'
        msg = 'There was an error loading the program\'s inputs which includes the Centricity file, OutputLog file, and appointment locations file.'
        emaillog(title, msg)
        sys.exit(0)

    # OUTPUT(site, date) Define and format date & time from locations book. Includes sanity checks for 'No Screening' locations, deletes unnecessary text from the location string and counts number of entries left in locations book.
    today = datetime.datetime.now()
    today = today.replace(hour=0, minute=0, second=0, microsecond=0)
    if rerun is True:
        nextapptdate = today + datetime.timedelta(days=1)
    else:
        nextapptdate = today + datetime.timedelta(days=2)
    y = -1
    for sheet in sheet_names:
        locationssheet = locationsbook.sheet_by_name(sheet_names[y])
        for row in range(locationssheet.nrows):
            val = locationssheet.cell(row, 1).value
            if (isinstance(val, float)):
                eventdate = datetime.datetime(*xlrd.xldate_as_tuple(val, locationsbook.datemode))
                if nextapptdate == eventdate:
                    date = eventdate
                    site = locationssheet.cell(row, 5).value
                    if site in noscreen:
                        logging.info('No screening on %s b/c the site is in the site exceptions list (i.e. noscreen) in references', eventdate)
                        print('No screening on %s b/c the site is in the site exceptions list (i.e. noscreen) in references', eventdate)
                        sys.exit(0)
                    for text in deltext:
                        site = site.replace(text, '')
                    if y == -1:
                        locations_left = locationssheet.nrows - row
                        if locations_left <= 10:
                            title = 'ACTION REQUIRED: Error in the Automated Text Reminder System!!'
                            msg = 'The calendar locations workbook needs to be updated. There are only a few locations left.'
                            emaillog(title, msg)
                            # break
        y = y - 1
    # # Verify the date and site variables are not blank.
    try:
        date
        site
    except:
        logging.info('No screening on %s b/c date or site were blank', nextapptdate)
        sys.exit(0)

    # Formated date to a list then remove the time portion or '00:00:00'
    date = list(str(date))
    del date[-9:]
    eventdate = str(''.join(date))
    date = eventdate

    # Log & return variables
    logging.info('Site for the screening on %s is %s', date, site)

    # Finding Centricity File(s) and setup for json file assembly
    os.chdir(centricity)
    files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
    os.chdir(script_dir)

    apptdate = datetime.datetime.now()
    if rerun == True:
        apptdate = apptdate + datetime.timedelta(days=-1)

    #this date is used to identify the correct file
    nextapptdate = apptdate.strftime("%Y%m%d")


    #change appt date to two days in the future to match the Centricity report time line
    apptdate = apptdate + datetime.timedelta(days=2)
    #this date is used to verify the date in file is accurate
    checkdate = apptdate.strftime("%Y-%m-%d")

    #make a copy of the files to remove without interfering with the possibility of messing up the loop.
    rmfiles = copy.deepcopy(files)
    #create a timestamp to enter into the Output log
    remindersent = datetime.datetime.now().strftime("%I:%M%p on %B %d, %Y")
    records = []
    lastloop = []
    partialphone = False

    print('Assembling records')

    for file in rmfiles:
        if len(rmfiles) > 10:
            try:
                os.remove(os.path.join(centricity, rmfiles[0]))
                del rmfiles[0]
            except:
                logging.critical('file not found', exc_info=True)

    for file in reversed(files):
        try:
            if nextapptdate in file:
                if 'COA' in file:
                    cat = ccat
                    creport = []
                    creport.append('Corporate Screening Report')
                    creport.append(0)
                    logging.info('%s file identified for Corporate Screenings', file)
                else:
                    cat = ocat
                    oreport = []
                    oreport.append('Outreach Screening Report')
                    oreport.append(0)
                    logging.info('%s File identified for Outreach Screenings', file)
                sourcebook = xlrd.open_workbook(os.path.join(centricity,file))
                sourcesheet = sourcebook.sheet_by_index(0)
                for rownum in range(sourcesheet.nrows):
                    if sourcesheet.cell_value(rownum, 0) not in exceptions:
                        IDnumber = IDnumber + 1
                        record = {
                            'date': date,
                            'location': site.title(),
                            'record_id': str(int(IDnumber)),
                            'remindersent': remindersent
                            }
                        for field in cat:
                            if field == 'phone':
                                logging.debug('inside the phone processor checking column %s', str(cat[field]))
                                if sourcesheet.cell(rownum, cat[field]).value != "":
                                    phone = sourcesheet.cell(rownum, cat[field]).value
                                    logging.debug('Raw phone number is %s', str(phone))
                                    phone = list(str(phone))
                                    del phone[0]
                                    del phone[3]
                                    phone = str(''.join(phone))
                                    record[field] = phone
                                    logging.debug('Formated phone number is %s', str(phone))

                            if field == 'time':
                                logging.debug('inside the time processor checking column %s', str(cat[field]))
                                sourcetime = sourcesheet.cell(rownum, cat[field]).value
                                time = datetime.datetime(*xlrd.xldate_as_tuple(sourcetime, sourcebook.datemode))
                                logging.debug('Raw time is %s', time)
                                time = list(str(time))
                                rec_date = str(''.join(time[:10]))

                                if rec_date != checkdate:
                                    logging.critical('the date found in the Centricity report does not match location')
                                    title = 'ACTION REQUIRED: Automated Appointment Reminder System Failure'
                                    msg = 'the date in the Centricity Report do not match the appointment location date'
                                    emaillog(title, msg)
                                    sys.exit(0)
                                del time[0:11]
                                del time[-3:]
                                time = str(''.join(time))
                                record[field] = time
                                logging.debug('Formatted time is %s', time)

                            elif field == 'fname' or field == 'lname':
                                logging.debug('inside everything else processor')
                                s = sourcesheet.cell_value(rownum, cat[field])
                                record[field] = s.title()

                        records.append(record)
                        if cat == ccat:
                            creport.append(len(records))
                        else:
                            oreport.append(len(records))

        except:
            logging.critical('The program failed in the assemble records function', exc_info=True)
            title = 'ACTION REQUIRED: Automated Reminder Program Error'
            msg = 'The loop that assembles records has failed.'
            emaillog(title, msg)
            sys.exit(0)

    try:
        logging.info('Records were successfully assembled')
        logging.info(oreport[0] + ' had ' + str(oreport[-1]) + ' records')
        logging.info(creport[0] + ' had ' + str(creport[-1]) + ' records')
    except:
        logging.critical('The program did not find both Centricity files in the loop', exc_info=True)
        title='ACTION REQUIRED: Automated Reminder Program Error'
        msg='Their was an issue logging the number of records imported from each file.'
        emaillog(title, msg)

    print('Logging records')
    try:
        logpath = os.path.join(logs, outputlog)
        logbook = openpyxl.load_workbook(logpath)
        logsheet = logbook.active
        firstrow = logsheet.max_row + 1
    except:
        logging.error('There was an error loading the Reminders log workbook')
        title = 'ACTION REQUIRED: Critical Error in the Appointment Reminder'
        msg = 'The Reminders log workbook failed to open.'
        emaillog(title, msg)
        sys.exit(0)
    # Taking advantage of the implicit booleanness of the empty list is very pythonic. This is a sanity check to ensure their are actual records in the list of records.
    if not records:
        logging.error('Records include %s. Program did not find any records to log.', records)
        title = 'ACTION REQUIRED: Potential Centricity Report Error'
        msg = 'The records do not contain any patient information. Check the Centricity report & make sure there is a screening on that date.'
        emaillog(title, msg)
        sys.exit(0)
    # in the list of records, in the key/value of each dictionary item, if the key (i.e. data label) matches that of logorder, the value of the record dictionary is placed in the .............
    for record in records:
        for k, v in record.items():
            if k in logorder:
                logsheet.cell(row = firstrow, column = logorder[k]).value = v
        firstrow += 1

    logbook.save(logpath)
    logging.info('Records were successfully logged in the Output log')
    return records

def send_records(records):
    print('Importing data into RedCap')
    if records:
        for record in records:
            del record['remindersent']
            record['patient_list_complete'] = '2'
        try:
            with open('api_keys.txt', 'r') as f:
                for i, line in enumerate(f):
                    if i == 1:
                        api = line
            api = api[:-1]
            URL = 'https://redcap.seton.org/redcap/api/'
            API_KEY = api
            project = redcap.Project(URL, API_KEY)
            submit = project.import_records(records)
        except:
            title = 'ACTION REQUIRED: Appointment Reminder Unable To Upload Records'
            msg = 'There was an error while trying to import records into redcap'
            emaillog(title, msg)
            logging.error('Error in the send_reminders function of the API')
            sys.exit(0)

        logging.info('Records were successfully imported to Redcap')
        print('All done!')
        tm.sleep(2)

if __name__ == '__main__':

    if not sys.argv[1:]:
        records = one()
    else:
        records = one(True)
    send_records(records)



